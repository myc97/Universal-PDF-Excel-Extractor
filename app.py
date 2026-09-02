```python
# ============================================================
# NEET UNIVERSAL PDF → EXCEL EXTRACTOR
# ============================================================
#
# Extraction priority:
#
# 1. Searchable PDF
#       ├── PyMuPDF text blocks
#       ├── Proven fixed-position extraction
#       ├── Header-position adaptive extraction
#       ├── Coordinate/X-position extraction
#       └── pdfplumber table extraction
#
# 2. Scanned PDF
#       └── OCR + adaptive row extraction
#
# Features:
#   - Different counselling PDF formats
#   - Non-table text PDFs
#   - Table PDFs
#   - Scanned PDFs
#   - College heading detection
#   - College carry-forward
#   - Automatic column-position detection
#   - Duplicate removal
#   - Excel output
#   - Error log
#   - Streamlit Cloud safe
#
# ============================================================

import gc
import io
import re
import tempfile
import time
from pathlib import Path

import fitz
import pandas as pd
import pdfplumber
import pytesseract
import streamlit as st

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from pdf2image import convert_from_path


# ============================================================
# STREAMLIT CONFIG
# ============================================================

st.set_page_config(
    page_title="NEET Universal PDF Extractor",
    page_icon="📄",
    layout="wide"
)


# ============================================================
# OUTPUT COLUMNS
# ============================================================

COLUMNS = [
    "Sr. No.",
    "AIR",
    "NEET Roll No.",
    "CET Form No.",
    "Name",
    "G",
    "Cat.",
    "Quota",
    "Code",
    "College",
    "PDF Page"
]


# ============================================================
# KNOWN POSITION FORMAT
#
# This is your proven format.
# It remains the FIRST fixed-position method.
# ============================================================

DEFAULT_POS = {
    "name": 39,
    "gender": 73,
    "cat": 76,
    "quota": 88,
    "code": 115,
    "college": 120
}


# ============================================================
# LIMITS
# ============================================================

MAX_BATCH = 100


# ============================================================
# TEXT CLEANING
# ============================================================

def clean(value):
    if value is None:
        return ""

    value = str(value)

    value = (
        value
        .replace("\xa0", " ")
        .replace("\r", " ")
        .replace("\n", " ")
    )

    value = re.sub(r"[ \t]+", " ", value)

    return value.strip()


def clean_college(value):
    value = clean(value)

    value = re.sub(
        r"^(college\s*[:\-]\s*)",
        "",
        value,
        flags=re.I
    )

    return value.strip()


# ============================================================
# NORMALIZE HEADER
# ============================================================

def normalize_header(value):

    value = clean(value).lower()

    value = re.sub(
        r"[^a-z0-9]+",
        " ",
        value
    ).strip()

    replacements = {
        "sr no": "sr",
        "sr": "sr",
        "serial no": "sr",
        "serial number": "sr",
        "s no": "sr",

        "air": "air",

        "neet roll no": "roll",
        "neet roll number": "roll",
        "roll no": "roll",
        "roll number": "roll",

        "cet form no": "form",
        "cet form number": "form",
        "form no": "form",
        "form number": "form",

        "candidate name": "name",
        "student name": "name",
        "name": "name",

        "gender": "gender",
        "sex": "gender",

        "category": "cat",
        "cat": "cat",

        "quota": "quota",

        "college code": "code",
        "institute code": "code",
        "code": "code",

        "college": "college",
        "college name": "college",
        "institute": "college",
        "institute name": "college"
    }

    return replacements.get(value, value)


# ============================================================
# HEADER DETECTION
# ============================================================

def header_line(text):

    if not text:
        return False

    low = re.sub(
        r"[^a-z]+",
        " ",
        text.lower()
    )

    keys = [
        "sr no",
        "air",
        "neet roll",
        "cet form",
        "name",
        "quota",
        "code",
        "college"
    ]

    score = sum(
        1 for key in keys
        if key in low
    )

    return score >= 3


# ============================================================
# DATA ROW DETECTION
# ============================================================

def first4(text):

    if not text:
        return None

    return re.match(
        r"^\s*"
        r"(?P<sr>\d+)\s+"
        r"(?P<air>\d+)\s+"
        r"(?P<roll>\d{8,12})\s+"
        r"(?P<form>\d{7,12})\s+",
        text
    )


def data_line(text):

    return first4(text) is not None


# ============================================================
# COLLEGE HEADING DETECTION
# ============================================================

def college_heading(text):

    text = clean(text)

    if not text:
        return False

    if data_line(text):
        return False

    if header_line(text):
        return False

    if len(text) < 6:
        return False

    low = text.lower()

    ignored = {
        "college list",
        "college wise",
        "allotment list",
        "seat allotment",
        "merit list",
        "candidate list",
        "result",
        "results",
        "instructions",
        "page"
    }

    if low in ignored:
        return False

    # Common college/institute indicators
    cues = [
        "medical college",
        "dental college",
        "college",
        "institute",
        "university",
        "hospital",
        "society",
        "academy",
        "medical sciences",
        "institute of medical",
        "medical science"
    ]

    if any(cue in low for cue in cues):
        return True

    # College code heading:
    # 1103 - ABC Medical College
    # 1103: ABC Medical College
    if re.match(
        r"^\s*\d{4}\s*[-:]\s*.+",
        text
    ):
        return True

    return False


# ============================================================
# EXTRACT COLLEGE HEADING
# ============================================================

def find_college_heading(lines, header_index=None):

    if not lines:
        return ""

    if header_index is None:
        header_index = len(lines)

    start = max(
        0,
        header_index - 10
    )

    candidates = []

    for line in lines[start:header_index]:

        value = clean(line)

        if college_heading(value):
            candidates.append(value)

    return (
        candidates[-1]
        if candidates
        else ""
    )


# ============================================================
# FITZ TEXT LINES
# ============================================================

def fitz_lines(page):

    lines = []

    blocks = page.get_text(
        "blocks",
        sort=True
    )

    for block in blocks:

        if len(block) < 5:
            continue

        text = block[4]

        if not text:
            continue

        for line in text.splitlines():

            line = line.rstrip(
                "\r\n"
            )

            if line.strip():
                lines.append(line)

    return lines


# ============================================================
# FITZ WORDS
# ============================================================

def fitz_words(page):

    try:
        words = page.get_text(
            "words",
            sort=True
        )

        return words or []

    except Exception:
        return []


# ============================================================
# FIXED POSITION ROW
#
# This preserves the method from your original working script.
# ============================================================

def fixed_row(
    line,
    page_no,
    college="",
    positions=None
):

    if not data_line(line):
        return None

    match = first4(line)

    if not match:
        return None

    pos = positions or DEFAULT_POS

    raw = line.rstrip(
        "\r\n"
    )

    college_start = pos["college"]

    if len(raw) < college_start:
        raw = raw.ljust(
            college_start
        )

    name = raw[
        pos["name"]:
        pos["gender"]
    ]

    gender = raw[
        pos["gender"]:
        pos["cat"]
    ]

    category = raw[
        pos["cat"]:
        pos["quota"]
    ]

    quota = raw[
        pos["quota"]:
        pos["code"]
    ]

    code = raw[
        pos["code"]:
        pos["college"]
    ]

    row_college = raw[
        pos["college"]:
    ]

    code = code.rstrip(
        ":"
    ).strip()

    code_match = re.search(
        r"\b\d{4}\b",
        code
    )

    if not code_match:
        return None

    code = code_match.group()

    row_college = clean_college(
        row_college
    )

    return {
        "Sr. No.": match.group("sr"),
        "AIR": match.group("air"),
        "NEET Roll No.": match.group("roll"),
        "CET Form No.": match.group("form"),
        "Name": clean(name),
        "G": clean(gender),
        "Cat.": clean(category),
        "Quota": clean(quota),
        "Code": code,
        "College": row_college or clean_college(college),
        "PDF Page": page_no
    }


# ============================================================
# HEADER POSITION DETECTION
#
# Used when the PDF has the same conceptual columns but
# different character spacing.
# ============================================================

def detect_positions_from_header(lines):

    header = None

    for line in lines:

        if header_line(line):
            header = line
            break

    if not header:
        return None

    positions = {}

    patterns = {
        "name": [
            r"\bName\b",
            r"\bCandidate\s+Name\b",
            r"\bStudent\s+Name\b"
        ],

        "gender": [
            r"\bG\b",
            r"\bGender\b",
            r"\bSex\b"
        ],

        "cat": [
            r"\bCat\.?\b",
            r"\bCategory\b"
        ],

        "quota": [
            r"\bQuota\b"
        ],

        "code": [
            r"\bCode\b",
            r"\bCollege\s+Code\b",
            r"\bInstitute\s+Code\b"
        ],

        "college": [
            r"\bCollege\b",
            r"\bCollege\s+Name\b",
            r"\bInstitute\b",
            r"\bInstitute\s+Name\b"
        ]
    }

    for key, plist in patterns.items():

        found = None

        for pattern in plist:

            m = re.search(
                pattern,
                header,
                flags=re.I
            )

            if m:
                found = m.start()
                break

        if found is not None:
            positions[key] = found

    required = [
        "name",
        "gender",
        "cat",
        "quota",
        "code",
        "college"
    ]

    if not all(
        key in positions
        for key in required
    ):
        return None

    # Sanity check:
    # columns should appear in left-to-right order.
    values = [
        positions[x]
        for x in required
    ]

    if values != sorted(values):
        return None

    return positions


# ============================================================
# HYBRID POSITION DETECTION
# ============================================================

def get_positions(lines):

    adaptive = detect_positions_from_header(
        lines
    )

    if adaptive:

        # Sometimes "G" is detected incorrectly because
        # single-letter matching is unsafe.
        # Only use adaptive result if spacing is sensible.

        if (
            adaptive["name"] < adaptive["gender"]
            < adaptive["cat"]
            < adaptive["quota"]
            < adaptive["code"]
            < adaptive["college"]
        ):
            return adaptive

    return DEFAULT_POS


# ============================================================
# COORDINATE BASED EXTRACTION
#
# This is the important fallback for PDFs where the textual
# spacing is destroyed but PDF coordinates remain correct.
# ============================================================

def coordinate_rows(
    page,
    page_no,
    college=""
):

    words = fitz_words(page)

    if not words:
        return []

    # Group words into visual lines.
    line_groups = []

    tolerance = 3.0

    for word in words:

        if len(word) < 5:
            continue

        x0, y0, x1, y1, text = word[:5]

        placed = False

        for group in line_groups:

            if abs(
                group["y"] - y0
            ) <= tolerance:

                group["words"].append(
                    word
                )

                # update average y
                group["y"] = (
                    group["y"] *
                    (len(group["words"]) - 1)
                    + y0
                ) / len(group["words"])

                placed = True
                break

        if not placed:

            line_groups.append(
                {
                    "y": y0,
                    "words": [word]
                }
            )

    line_groups.sort(
        key=lambda x: x["y"]
    )

    records = []

    current_college = college

    for group in line_groups:

        row_words = sorted(
            group["words"],
            key=lambda w: w[0]
        )

        text = " ".join(
            clean(w[4])
            for w in row_words
        )

        if college_heading(text):
            current_college = text
            continue

        if not data_line(text):
            continue

        match = first4(text)

        if not match:
            continue

        # Find the four leading numeric words.
        numeric_indices = []

        for i, word in enumerate(row_words):

            value = clean(word[4])

            if re.fullmatch(
                r"\d+",
                value
            ):
                numeric_indices.append(i)

        if len(numeric_indices) < 4:
            continue

        # The first four numeric fields are expected
        # to be Sr / AIR / Roll / Form.
        first_indices = numeric_indices[:4]

        if first_indices != [0, 1, 2, 3]:
            # Some PDFs have an empty/merged word.
            # Continue with regex method.
            pass

        # Locate 4-digit college code.
        code_index = None

        for i, word in enumerate(row_words):

            value = clean(word[4])

            if re.fullmatch(
                r"\d{4}:?",
                value
            ):
                code_index = i
                break

        if code_index is None:
            continue

        # Text between Form and Code
        body_words = row_words[4:code_index]

        body = [
            clean(w[4])
            for w in body_words
        ]

        body = [
            x for x in body
            if x
        ]

        if len(body) < 4:
            continue

        # Best-effort separation:
        #
        # Name | Gender | Category | Quota
        #
        # Gender usually one/two characters.
        # Category generally short.
        # Quota may contain spaces.

        gender_index = None
        cat_index = None

        for i, value in enumerate(body):

            if re.fullmatch(
                r"[MF]",
                value,
                flags=re.I
            ):
                gender_index = i
                break

        if gender_index is not None:

            name = " ".join(
                body[:gender_index]
            )

            gender = body[
                gender_index
            ]

            remaining = body[
                gender_index + 1:
            ]

            if remaining:

                cat_index = 0

                category = remaining[0]

                quota = " ".join(
                    remaining[1:]
                )

            else:
                category = ""
                quota = ""

        else:

            # Conservative fallback
            name = body[0]
            gender = ""
            category = ""
            quota = " ".join(
                body[1:]
            )

        after_code = row_words[
            code_index + 1:
        ]

        row_college = clean_college(
            " ".join(
                clean(w[4])
                for w in after_code
            )
        )

        code = clean(
            row_words[
                code_index
            ][4]
        ).rstrip(":")

        records.append(
            {
                "Sr. No.": match.group("sr"),
                "AIR": match.group("air"),
                "NEET Roll No.": match.group("roll"),
                "CET Form No.": match.group("form"),
                "Name": clean(name),
                "G": clean(gender),
                "Cat.": clean(category),
                "Quota": clean(quota),
                "Code": code,
                "College": (
                    row_college
                    or clean_college(
                        current_college
                    )
                ),
                "PDF Page": page_no
            }
        )

    return records


# ============================================================
# PDFPLUMBER TABLE EXTRACTION
# ============================================================

def table_records(
    table,
    page_no,
    college=""
):

    rows = []

    for row in table or []:

        if not row:
            continue

        cleaned = [
            clean(x)
            for x in row
        ]

        if any(cleaned):
            rows.append(cleaned)

    if not rows:
        return []

    aliases = {
        "sr",
        "air",
        "roll",
        "form",
        "name",
        "gender",
        "cat",
        "quota",
        "code",
        "college"
    }

    header_index = -1
    best_score = 0

    for i, row in enumerate(
        rows[:10]
    ):

        score = sum(
            normalize_header(x)
            in aliases
            for x in row
        )

        if score > best_score:
            best_score = score
            header_index = i

    if header_index < 0:
        return []

    if best_score < 3:
        return []

    mapping = {}

    for i, value in enumerate(
        rows[header_index]
    ):

        key = normalize_header(
            value
        )

        if (
            key in aliases
            and key not in mapping
        ):
            mapping[key] = i

    required = {
        "sr",
        "air",
        "roll",
        "form",
        "name",
        "quota",
        "code"
    }

    if not required.issubset(
        mapping
    ):
        return []

    output = []

    for row in rows[
        header_index + 1:
    ]:

        def val(key):

            index = mapping.get(key)

            if (
                index is None
                or index >= len(row)
            ):
                return ""

            return clean(
                row[index]
            )

        if not re.fullmatch(
            r"\d+",
            val("sr")
        ):
            continue

        if not re.fullmatch(
            r"\d+",
            val("air")
        ):
            continue

        if not re.fullmatch(
            r"\d{8,12}",
            val("roll")
        ):
            continue

        if not re.fullmatch(
            r"\d{7,12}",
            val("form")
        ):
            continue

        code_match = re.search(
            r"\b\d{4}\b",
            val("code")
        )

        if not code_match:
            continue

        output.append(
            {
                "Sr. No.": val("sr"),
                "AIR": val("air"),
                "NEET Roll No.": val("roll"),
                "CET Form No.": val("form"),
                "Name": val("name"),
                "G": val("gender"),
                "Cat.": val("cat"),
                "Quota": val("quota"),
                "Code": code_match.group(),
                "College": (
                    clean_college(
                        val("college")
                    )
                    or clean_college(
                        college
                    )
                ),
                "PDF Page": page_no
            }
        )

    return output


# ============================================================
# SEARCHABLE PAGE EXTRACTION
# ============================================================

def searchable_page(
    page,
    page_no,
    current_college=""
):

    lines = fitz_lines(page)

    if not lines:
        return [], current_college

    # --------------------------------------------------------
    # Detect college heading
    # --------------------------------------------------------

    header_index = next(
        (
            i
            for i, line in enumerate(lines)
            if header_line(line)
        ),
        None
    )

    heading = find_college_heading(
        lines,
        header_index
    )

    if heading:
        current_college = clean_college(
            heading
        )

    # --------------------------------------------------------
    # METHOD 1
    # Proven fixed-position extraction
    # --------------------------------------------------------

    positions = get_positions(
        lines
    )

    fixed_records = []

    for line in lines:

        row = fixed_row(
            line,
            page_no,
            current_college,
            positions
        )

        if row:

            # Avoid obviously invalid rows
            if (
                row["Name"]
                or row["College"]
            ):
                fixed_records.append(
                    row
                )

    if fixed_records:

        # Carry college from successful rows
        for row in fixed_records:

            if not row["College"]:
                row["College"] = (
                    current_college
                )

            elif not current_college:
                current_college = (
                    row["College"]
                )

        return (
            fixed_records,
            current_college
        )

    # --------------------------------------------------------
    # METHOD 2
    # Coordinate extraction
    # --------------------------------------------------------

    coordinate_records = coordinate_rows(
        page,
        page_no,
        current_college
    )

    if coordinate_records:

        for row in coordinate_records:

            if row["College"]:
                current_college = (
                    row["College"]
                )

        return (
            coordinate_records,
            current_college
        )

    # --------------------------------------------------------
    # METHOD 3
    # pdfplumber tables
    # --------------------------------------------------------

    table_records_all = []

    try:

        tables = page.extract_tables()

        for table in tables or []:

            table_records_all.extend(
                table_records(
                    table,
                    page_no,
                    current_college
                )
            )

    except Exception:
        pass

    if table_records_all:

        for row in table_records_all:

            if row["College"]:
                current_college = (
                    row["College"]
                )

        return (
            table_records_all,
            current_college
        )

    return [], current_college


# ============================================================
# OCR ROW EXTRACTION
# ============================================================

def ocr_extract_lines(
    lines,
    page_no,
    current_college=""
):

    records = []

    header_index = next(
        (
            i
            for i, line in enumerate(lines)
            if header_line(line)
        ),
        None
    )

    heading = find_college_heading(
        lines,
        header_index
    )

    if heading:
        current_college = clean_college(
            heading
        )

    # --------------------------------------------------------
    # First attempt:
    # fixed-position method
    # --------------------------------------------------------

    positions = get_positions(
        lines
    )

    for line in lines:

        row = fixed_row(
            line,
            page_no,
            current_college,
            positions
        )

        if row:

            records.append(row)

    if records:

        for row in records:

            if row["College"]:
                current_college = (
                    row["College"]
                )

        return records, current_college

    # --------------------------------------------------------
    # OCR conservative fallback
    #
    # Keep first four numeric fields.
    # Find four-digit college code.
    # --------------------------------------------------------

    for line in lines:

        match = first4(line)

        if not match:
            continue

        tail = line[
            match.end():
        ]

        code_match = re.search(
            r"\b\d{4}\b",
            tail
        )

        if not code_match:
            continue

        before = tail[
            :code_match.start()
        ]

        after = tail[
            code_match.end():
        ]

        # OCR frequently uses multiple spaces
        parts = [
            clean(x)
            for x in re.split(
                r"\s{2,}",
                before
            )
            if clean(x)
        ]

        if len(parts) < 3:
            continue

        # Try to identify:
        #
        # Name / Gender / Category / Quota

        name = ""
        gender = ""
        category = ""
        quota = ""

        if len(parts) >= 4:

            name = parts[0]

            if re.fullmatch(
                r"[MF]",
                parts[1],
                flags=re.I
            ):
                gender = parts[1]
                category = parts[2]
                quota = " ".join(
                    parts[3:]
                )

            else:

                gender = parts[-3]
                category = parts[-2]
                quota = parts[-1]

        else:

            name = parts[0]
            quota = parts[-1]

        row_college = clean_college(
            after
        )

        row = {
            "Sr. No.": match.group("sr"),
            "AIR": match.group("air"),
            "NEET Roll No.": match.group("roll"),
            "CET Form No.": match.group("form"),
            "Name": clean(name),
            "G": clean(gender),
            "Cat.": clean(category),
            "Quota": clean(quota),
            "Code": code_match.group(),
            "College": (
                row_college
                or clean_college(
                    current_college
                )
            ),
            "PDF Page": page_no
        }

        records.append(row)

        if row["College"]:
            current_college = row["College"]

    return records, current_college


# ============================================================
# OCR ONE PAGE
# ============================================================

def ocr_page(
    pdf_path,
    page_no,
    dpi,
    current_college=""
):

    img = None

    try:

        images = convert_from_path(
            pdf_path,
            dpi=dpi,
            first_page=page_no,
            last_page=page_no,
            use_cropbox=True,
            thread_count=1
        )

        if not images:
            return [], current_college

        img = images[0]

        # PSM 6 works well for counselling lists.
        text = pytesseract.image_to_string(
            img,
            lang="eng",
            config="--psm 6"
        )

        lines = [
            x
            for x in text.splitlines()
            if x.strip()
        ]

        return ocr_extract_lines(
            lines,
            page_no,
            current_college
        )

    except Exception:
        return [], current_college

    finally:

        try:
            if img:
                img.close()
        except Exception:
            pass

        gc.collect()


# ============================================================
# PDF TYPE DETECTION
# ============================================================

def detect_type(path):

    try:

        with fitz.open(path) as doc:

            sample_pages = min(
                5,
                len(doc)
            )

            text_length = 0

            for i in range(
                sample_pages
            ):

                text = (
                    doc[i].get_text(
                        "text"
                    )
                    or ""
                )

                text_length += len(
                    text.strip()
                )

            if text_length >= 50:
                return "searchable"

    except Exception:
        pass

    return "scanned"


# ============================================================
# QUALITY CHECK
# ============================================================

def row_quality(row):

    score = 0

    if re.fullmatch(
        r"\d+",
        str(row.get("Sr. No.", ""))
    ):
        score += 1

    if re.fullmatch(
        r"\d+",
        str(row.get("AIR", ""))
    ):
        score += 1

    if re.fullmatch(
        r"\d{8,12}",
        str(row.get("NEET Roll No.", ""))
    ):
        score += 1

    if re.fullmatch(
        r"\d{7,12}",
        str(row.get("CET Form No.", ""))
    ):
        score += 1

    if row.get("Name"):
        score += 1

    if row.get("Quota"):
        score += 1

    if re.fullmatch(
        r"\d{4}",
        str(row.get("Code", ""))
    ):
        score += 2

    return score


# ============================================================
# NORMALIZE DATA
# ============================================================

def normalize(records):

    if not records:
        return None, 0

    df = pd.DataFrame(
        records
    )

    # Ensure every output column exists
    for column in COLUMNS:

        if column not in df.columns:
            df[column] = ""

    df = df[
        COLUMNS
    ].copy()

    for column in COLUMNS:

        df[column] = (
            df[column]
            .map(clean)
        )

    # --------------------------------------------------------
    # Remove rows that are clearly invalid
    # --------------------------------------------------------

    df["_quality"] = df.apply(
        row_quality,
        axis=1
    )

    df = df[
        df["_quality"] >= 5
    ].copy()

    df.drop(
        columns=["_quality"],
        inplace=True
    )

    # --------------------------------------------------------
    # College recovery by Code
    # --------------------------------------------------------

    code_map = {}

    for _, row in df.iterrows():

        code = row["Code"]
        college = row["College"]

        if code and college:

            code_map.setdefault(
                code,
                college
            )

    missing_college = (
        df["College"].eq("")
    )

    df.loc[
        missing_college,
        "College"
    ] = (
        df.loc[
            missing_college,
            "Code"
        ]
        .map(code_map)
        .fillna("")
    )

    # --------------------------------------------------------
    # Remove duplicates
    # --------------------------------------------------------

    before = len(df)

    duplicate_subset = [
        "Sr. No.",
        "AIR",
        "NEET Roll No.",
        "CET Form No.",
        "Name",
        "G",
        "Cat.",
        "Quota",
        "Code",
        "College"
    ]

    df = df.drop_duplicates(
        subset=duplicate_subset,
        keep="first"
    )

    duplicates = (
        before - len(df)
    )

    # --------------------------------------------------------
    # Sort
    # --------------------------------------------------------

    df["_page"] = pd.to_numeric(
        df["PDF Page"],
        errors="coerce"
    )

    df["_sr"] = pd.to_numeric(
        df["Sr. No."],
        errors="coerce"
    )

    df = (
        df
        .sort_values(
            [
                "_page",
                "_sr"
            ],
            na_position="last"
        )
        .drop(
            columns=[
                "_page",
                "_sr"
            ]
        )
        .reset_index(
            drop=True
        )
    )

    return df, duplicates


# ============================================================
# EXCEL GENERATION
# ============================================================

def excel_bytes(
    df,
    stats
):

    buffer = io.BytesIO()

    with pd.ExcelWriter(
        buffer,
        engine="openpyxl"
    ) as writer:

        # Excel maximum rows is 1,048,576.
        # Keep a safe limit.
        sheet_limit = 1_048_000

        for number, start in enumerate(
            range(
                0,
                len(df),
                sheet_limit
            ),
            1
        ):

            sheet_name = (
                "Extracted_Data"
                if number == 1
                else f"Data_{number}"
            )

            df.iloc[
                start:
                start + sheet_limit
            ].to_excel(
                writer,
                index=False,
                sheet_name=sheet_name
            )

        pd.DataFrame(
            [stats]
        ).to_excel(
            writer,
            index=False,
            sheet_name="Extraction_Summary"
        )

    buffer.seek(0)

    wb = load_workbook(
        buffer
    )

    # --------------------------------------------------------
    # Format worksheets
    # --------------------------------------------------------

    for ws in wb.worksheets:

        ws.freeze_panes = "A2"

        if ws.max_row > 1:
            ws.auto_filter.ref = (
                ws.dimensions
            )

        for cell in ws[1]:

            cell.font = Font(
                bold=True
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

    # --------------------------------------------------------
    # Data column widths
    # --------------------------------------------------------

    widths = [
        12,  # Sr No
        12,  # AIR
        20,  # Roll
        20,  # Form
        38,  # Name
        8,   # Gender
        14,  # Category
        30,  # Quota
        10,  # Code
        60,  # College
        12   # Page
    ]

    for ws in wb.worksheets:

        if ws.title == "Extraction_Summary":
            continue

        for index, width in enumerate(
            widths,
            1
        ):

            column_letter = (
                chr(64 + index)
            )

            ws.column_dimensions[
                column_letter
            ].width = width

        for row in ws.iter_rows(
            min_row=2
        ):

            for cell in row:

                cell.alignment = Alignment(
                    vertical="center"
                )

    output = io.BytesIO()

    wb.save(output)

    return output.getvalue()


# ============================================================
# MAIN EXTRACTION ENGINE
# ============================================================

def run_extraction(
    path,
    batch_size,
    dpi,
    progress,
    status
):

    started = time.time()

    pdf_type = detect_type(
        path
    )

    with fitz.open(path) as doc:
        total_pages = len(doc)

    records = []
    errors = []

    current_college = ""

    batches = [
        list(
            range(
                start,
                min(
                    start + batch_size,
                    total_pages
                )
            )
        )
        for start in range(
            0,
            total_pages,
            batch_size
        )
    ]

    status.info(
        f"Detected **{pdf_type.title()} PDF** | "
        f"Pages: **{total_pages:,}** | "
        f"Batches: **{len(batches):,}**"
    )

    # ========================================================
    # SEARCHABLE PDF
    # ========================================================

    if pdf_type == "searchable":

        # Open once instead of opening pdfplumber
        # repeatedly for every batch.
        try:

            with pdfplumber.open(path) as pdf:

                for batch_number, indexes in enumerate(
                    batches,
                    1
                ):

                    batch_start = time.time()

                    for index in indexes:

                        page_no = index + 1

                        try:

                            page = pdf.pages[
                                index
                            ]

                            page_records, current_college = (
                                searchable_page(
                                    page,
                                    page_no,
                                    current_college
                                )
                            )

                            records.extend(
                                page_records
                            )

                        except Exception as page_error:

                            errors.append(
                                f"Page {page_no}: "
                                f"{page_error}"
                            )

                    elapsed = (
                        time.time()
                        - started
                    )

                    if batch_number:

                        eta = (
                            elapsed
                            / batch_number
                            * (
                                len(batches)
                                - batch_number
                            )
                        )

                    else:
                        eta = 0

                    progress.progress(
                        batch_number
                        / len(batches)
                    )

                    status.info(
                        f"Batch **{batch_number}/"
                        f"{len(batches)}** | "
                        f"Pages **{indexes[0] + 1}-"
                        f"{indexes[-1] + 1}** | "
                        f"Rows: **{len(records):,}** | "
                        f"Batch: **{time.time() - batch_start:.1f}s** | "
                        f"ETA: **{eta:.0f}s**"
                    )

                    gc.collect()

        except Exception as e:

            errors.append(
                f"PDF processing error: {e}"
            )

    # ========================================================
    # SCANNED PDF
    # ========================================================

    else:

        for batch_number, indexes in enumerate(
            batches,
            1
        ):

            batch_start = time.time()

            for index in indexes:

                page_no = index + 1

                try:

                    page_records, current_college = (
                        ocr_page(
                            path,
                            page_no,
                            dpi,
                            current_college
                        )
                    )

                    records.extend(
                        page_records
                    )

                except Exception as page_error:

                    errors.append(
                        f"Page {page_no}: "
                        f"{page_error}"
                    )

            elapsed = (
                time.time()
                - started
            )

            eta = (
                elapsed
                / batch_number
                * (
                    len(batches)
                    - batch_number
                )
            )

            progress.progress(
                batch_number
                / len(batches)
            )

            status.info(
                f"OCR Batch **{batch_number}/"
                f"{len(batches)}** | "
                f"Pages **{indexes[0] + 1}-"
                f"{indexes[-1] + 1}** | "
                f"Rows: **{len(records):,}** | "
                f"Batch: **{time.time() - batch_start:.1f}s** | "
                f"ETA: **{eta:.0f}s**"
            )

            gc.collect()

    # ========================================================
    # NORMALIZE
    # ========================================================

    df, duplicates = normalize(
        records
    )

    stats = {
        "PDF Type": pdf_type,
        "Pages Processed": total_pages,
        "Batches": len(batches),
        "Raw Rows Detected": len(records),
        "Rows Extracted": (
            0
            if df is None
            else len(df)
        ),
        "Duplicates Removed": duplicates,
        "Errors": len(errors),
        "Time Seconds": round(
            time.time() - started,
            2
        )
    }

    return (
        df,
        stats,
        errors
    )


# ============================================================
# STREAMLIT UI
# ============================================================

st.title(
    "📄 NEET Universal PDF → Excel Extractor"
)

st.caption(
    "PyMuPDF fixed-position + adaptive coordinates + "
    "table extraction + college detection + OCR"
)


# ============================================================
# SIDEBAR
# ============================================================

with st.sidebar:

    st.header(
        "⚙️ Settings"
    )

    batch_size = st.number_input(
        "Batch Size",
        min_value=1,
        max_value=MAX_BATCH,
        value=50,
        step=10
    )

    dpi = st.slider(
        "OCR DPI",
        min_value=120,
        max_value=300,
        value=200,
        step=10
    )

    st.subheader(
        "Output Columns"
    )

    st.code(
        "\n".join(COLUMNS)
    )

    st.markdown(
        """
### Extraction Priority

**Searchable PDF**
1. Fixed-position
2. Header adaptive
3. Coordinate based
4. pdfplumber table

**Scanned PDF**
1. OCR
2. Fixed-position
3. Conservative fallback
"""
    )


# ============================================================
# FILE UPLOAD
# ============================================================

uploaded = st.file_uploader(
    "Upload NEET Counselling / Allotment PDF",
    type=["pdf"]
)


# ============================================================
# PROCESS
# ============================================================

if uploaded:

    st.success(
        f"Loaded: **{uploaded.name}** "
        f"({uploaded.size / (1024 * 1024):.2f} MB)"
    )

    if st.button(
        "🚀 Start Extraction",
        type="primary",
        use_container_width=True
    ):

        temp_file = tempfile.NamedTemporaryFile(
            delete=False,
            suffix=".pdf"
        )

        temp_file.write(
            uploaded.getbuffer()
        )

        temp_file.close()

        progress = st.progress(
            0
        )

        status = st.empty()

        try:

            with st.spinner(
                "Processing PDF..."
            ):

                df, stats, errors = (
                    run_extraction(
                        temp_file.name,
                        int(batch_size),
                        int(dpi),
                        progress,
                        status
                    )
                )

            # ==================================================
            # NO DATA
            # ==================================================

            if (
                df is None
                or df.empty
            ):

                st.error(
                    "❌ No valid rows were detected."
                )

                st.warning(
                    "This PDF may use a structure that "
                    "requires OCR or a different column layout."
                )

                if errors:

                    with st.expander(
                        f"Processing errors ({len(errors)})"
                    ):

                        st.text(
                            "\n".join(
                                errors[:300]
                            )
                        )

            # ==================================================
            # SUCCESS
            # ==================================================

            else:

                st.success(
                    f"🎉 Extraction completed — "
                    f"**{len(df):,} rows** extracted."
                )

                a, b, c, d, e = st.columns(5)

                a.metric(
                    "Pages",
                    f"{stats['Pages Processed']:,}"
                )

                b.metric(
                    "Raw Rows",
                    f"{stats['Raw Rows Detected']:,}"
                )

                c.metric(
                    "Final Rows",
                    f"{stats['Rows Extracted']:,}"
                )

                d.metric(
                    "Duplicates",
                    f"{stats['Duplicates Removed']:,}"
                )

                e.metric(
                    "Errors",
                    f"{stats['Errors']:,}"
                )

                # ==================================================
                # PREVIEW
                # ==================================================

                st.subheader(
                    "📊 Extracted Data Preview"
                )

                st.dataframe(
                    df.head(100),
                    use_container_width=True,
                    height=500
                )

                # ==================================================
                # STATISTICS
                # ==================================================

                st.subheader(
                    "📈 Extraction Summary"
                )

                summary_df = pd.DataFrame(
                    [
                        {
                            "Metric": "PDF Type",
                            "Value": stats["PDF Type"]
                        },
                        {
                            "Metric": "Pages Processed",
                            "Value": stats["Pages Processed"]
                        },
                        {
                            "Metric": "Raw Rows",
                            "Value": stats["Raw Rows Detected"]
                        },
                        {
                            "Metric": "Final Rows",
                            "Value": stats["Rows Extracted"]
                        },
                        {
                            "Metric": "Duplicates Removed",
                            "Value": stats["Duplicates Removed"]
                        },
                        {
                            "Metric": "Errors",
                            "Value": stats["Errors"]
                        },
                        {
                            "Metric": "Processing Time",
                            "Value": f"{stats['Time Seconds']} seconds"
                        }
                    ]
                )

                st.dataframe(
                    summary_df,
                    hide_index=True,
                    use_container_width=True
                )

                # ==================================================
                # EXCEL
                # ==================================================

                excel_data = excel_bytes(
                    df,
                    stats
                )

                st.download_button(
                    "⬇️ Download Excel",
                    data=excel_data,
                    file_name=(
                        f"{Path(uploaded.name).stem}"
                        f"_Extracted.xlsx"
                    ),
                    mime=(
                        "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet"
                    ),
                    use_container_width=True
                )

                # ==================================================
                # ERRORS
                # ==================================================

                if errors:

                    with st.expander(
                        f"⚠️ Processing errors ({len(errors)})"
                    ):

                        st.text(
                            "\n".join(
                                errors[:300]
                            )
                        )

        except Exception as e:

            st.error(
                f"❌ Extraction failed: {e}"
            )

        finally:

            try:

                Path(
                    temp_file.name
                ).unlink(
                    missing_ok=True
                )

            except Exception:
                pass

            gc.collect()

else:

    st.info(
        "Upload a NEET counselling/allotment PDF to begin."
    )


# ============================================================
# END
# ============================================================
```
